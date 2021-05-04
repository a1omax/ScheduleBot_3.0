import openpyxl
import os
import sqlite3

from openpyxl.utils import rows_from_range

for FILENAME in os.listdir():
    if ".xlsx" in FILENAME:
        if "~$" not in FILENAME:
            break
else:
    raise NameError("File .xlsx not found")


con = sqlite3.connect('../schedule.db')
cur = con.cursor()
cur.execute("""CREATE TABLE IF NOT EXISTS schedule(
                    group_name TEXT,
                    week_type TEXT,
                    week_day TEXT,
                    number_para INT,
                    name_para_first TEXT,
                    cabinet_first TEXT,
                    name_para_second TEXT,
                    cabinet_second TEXT
                    )""")
cur.execute("""DROP TABLE IF EXISTS schedule""")
con.commit()

print(FILENAME)


def start_cell():

    for i in range(1, 20):
        if get_value([ws["A" + str(i)]])[0] == "День":
            global s_cell
            s_cell = i-5
            return 1

    raise NameError("Нет клетки День")


def read_wb(number_list):            # Чтение листа эксель
    global ws
    wb = openpyxl.load_workbook(filename=FILENAME)
    ws = wb[wb.sheetnames[number_list]]


# Принимает merged клетку и возвращает значение merged клетки
def merged_cell_value(cell):
    idx = cell.coordinate
    for range_ in ws.merged_cells.ranges:
        merged_cells = list(openpyxl.utils.rows_from_range(str(range_)))
        for row in merged_cells:
            if idx in row:
                # If this is a merged cell,
                # return  the first cell of the merge range
                return ws[merged_cells[0][0]].value


# Принимает диапазон по одной букве и возвращает массив клеток
def get_cell_array(col, start, finish, step):
    letter = openpyxl.utils.cell.get_column_letter(col)
    array_cells = []
    i = 0
    start += s_cell
    finish += s_cell
    while True:
        num_cell = start + step * i
        if num_cell > finish:
            break
        array_cells.append(ws[letter + str(num_cell)])
        i += 1
    return array_cells


# Принимает массив клеток и возвращает массив их значений (включая merged)
def get_value(cell_array, flag=True):

    value_cell_array = []

    for i in range(len(cell_array)):

        cell = cell_array[i]

        if isinstance(cell, openpyxl.cell.cell.Cell):
            value = cell.value
        elif isinstance(cell, openpyxl.cell.cell.MergedCell):
            if flag:
                value = merged_cell_value(cell)
            else:
                value = None
        else:
            raise NameError("Cell has a wrong type")

        if isinstance(value, str):
            value = value.replace("        ", " ")  # x8
            value = value.replace("       ", " ")  # x7
            value = value.replace("      ", " ")  # x6
            value = value.replace("     ", " ")  # x5
            value = value.replace("    ", " ")  # x4
            value = value.replace("   ", " ")  # x3

        value_cell_array.append(value)
    return value_cell_array


def get_group_names_array():           # Возвращает массив названий всех групп
    group_names = []
    i = 3
    while True:
        group = get_value(get_cell_array(i, 5, 5, 1))[0]

        if group != "День":

            group_names.append(group)  # group row
            i += 3
        else:
            break

    return group_names


def get_num_para_array():          # Возвращает массив номеров пар
    num_para_array = []
    num_array = get_cell_array(2, 6, 300, 4)
    for i in num_array:
        num_para = get_value([i])[0]
        if num_para is None:
            break
        else:
            num_para_array.append(num_para)

    return num_para_array


def get_schedule(group_names, num_paras):
    len_letters = len(group_names)
    len_numbers = len(num_paras)

    array_first_odd = []
    array_second_odd = []
    array_first_even = []
    array_second_even = []

    arr_cab_f_o = []
    arr_cab_s_o = []
    arr_cab_f_e = []
    arr_cab_s_e = []

    for letter in range(len_letters):

        array_first_odd.append(get_value(get_cell_array(
            (3 + 3 * letter), 6, len_numbers * 4 + 2, 4)))
        array_second_odd.append(get_value(get_cell_array(
            (4 + 3 * letter), 6, len_numbers * 4 + 2, 4)))
        array_first_even.append(get_value(get_cell_array(
            (3 + 3 * letter), 8, len_numbers * 4 + 4, 4)))
        array_second_even.append(get_value(get_cell_array(
            (4 + 3 * letter), 8, len_numbers * 4 + 4, 4)))

        arr_cab_f_o.append(get_value(get_cell_array(
            (5 + 3 * letter), 6, len_numbers * 4 + 2, 4), False))
        arr_cab_s_o.append(get_value(get_cell_array(
            (5 + 3 * letter), 7, len_numbers * 4 + 3, 4), False))
        arr_cab_f_e.append(get_value(get_cell_array(
            (5 + 3 * letter), 8, len_numbers * 4 + 4, 4), False))
        arr_cab_s_e.append(get_value(get_cell_array(
            (5 + 3 * letter), 9, len_numbers * 4 + 5, 4), False))

    for l in range(len_letters):
        day = 0
        for c in range(len_numbers):

            if arr_cab_f_e[l][c] is None:

                arr_cab_f_e[l][c] = arr_cab_f_o[l][c]

            if arr_cab_s_e[l][c] is None:

                arr_cab_s_e[l][c] = arr_cab_f_e[l][c]

            if arr_cab_s_o[l][c] is None:

                arr_cab_s_o[l][c] = arr_cab_f_o[l][c]

            if array_first_even[l][c] is None:
                arr_cab_f_e[l][c] = None

            if array_second_even[l][c] is None:
                arr_cab_s_e[l][c] = None

            if array_second_odd is None:
                arr_cab_s_o[l][c] = None

            if num_paras[c] == num_paras[0] and c != 0:
                day += 1

            con = sqlite3.connect('../schedule.db')
            cur = con.cursor()
            cur.execute("""CREATE TABLE IF NOT EXISTS schedule(
                    group_name TEXT,
                    week_type TEXT,
                    week_day TEXT,
                    number_para INT,
                    name_para_first TEXT,
                    cabinet_first TEXT,
                    name_para_second TEXT,
                    cabinet_second TEXT
                    )""")
            cur.execute("""INSERT INTO schedule VALUES (?,?,?,?,?,?,?,?)""", (group_names[l], "odd", day, num_paras[c],
                        array_first_odd[l][c], arr_cab_f_o[l][c], array_second_odd[l][c], arr_cab_s_o[l][c]))
            cur.execute("""INSERT INTO schedule VALUES (?,?,?,?,?,?,?,?)""", (group_names[l], "even", day, num_paras[c],
                                                                              array_first_even[l][c], arr_cab_f_e[l][c],
                                                                              array_second_even[l][c], arr_cab_s_e[l][c]))
            con.commit()

    return [[array_first_odd, array_second_odd], [array_first_even, array_second_even]], [[arr_cab_f_o, arr_cab_s_o], [arr_cab_f_e, arr_cab_s_e]]


def main():

    all_group_names = []
    all_num_paras = []
    all_schedules = []
    all_cabinets = []
    for ws_number in range(8):
        print(ws_number)
        read_wb(ws_number)                  # чтение 8 листов по очереди
        start_cell()
        # запись всех 8 массивов названий групп в один массив
        all_group_names.append(get_group_names_array())
        all_num_paras.append(get_num_para_array())

        schedule, cabinet = get_schedule(
            all_group_names[ws_number], all_num_paras[ws_number])
        all_schedules.append(schedule)
        all_cabinets.append(cabinet)

        print(all_group_names[ws_number])
        print(all_num_paras[ws_number])
        print(all_schedules[ws_number])
        print(all_cabinets[ws_number])

    print(all_group_names)
    print(all_num_paras)
    print(all_cabinets)

    return 0


if __name__ == "__main__":
    main()
